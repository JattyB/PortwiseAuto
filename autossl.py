#!/usr/bin/env python3

import subprocess
import os
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Color
import re

##install xlsxwrite -> pip3 install xlsxwriter

testssl = "/home/jatty/gitools/testssl.sh/testssl.sh" ## Change this!!

ip_l1 = []
prt_l1 = []
TC1  = [None]
TC2  = [None]
TC3  = [None]
TC4  = [None]
TC5  = [None]
TC6  = [None]
TC7  = [None]
TC8  = [None]
TC9  = [None]
TC10 = [None]
DM   = {}
dss = {}

ip_pattern = "(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})"

def TEST_CASE_1(csvfile):
	### Certificate Expired
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "CERT_EXPIRATIONSTATUS":
				if row['finding'].upper() == "EXPIRED":
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_2(csvfile):
	### SSLv3 POODLE
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "POODLE_SSL":
				if "NOT VULNERABLE" not in row['finding'].upper():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_3(csvfile):
	### TLS Server Supports TLS v1.0 and TLS v1.1 and Weak Cipher Algorithms
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "TLS1":
				if row['finding'].lower() == "offered (deprecated)" or row['finding'].lower() == "offered":
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")
			if row["id"].upper() == "TLS1_1":
				if row['finding'].lower() == "offered (deprecated)" or row['finding'].lower() == "offered":
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")


def TEST_CASE_4(csvfile):
	### Secure Client-Initiated Renegotiation
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "SECURE_CLIENT_RENEGO":
				if "NOT VULNERABLE" not in row['finding'].upper():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_5(csvfile):
	### HSTS Policy not implemented
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "HSTS":
				if "NOT OFFERED" in row['finding'].upper():
						return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_6(csvfile):
	### SSL/TLS Diffie-Hellman Modulus <= 1024 Bits (Logjam)
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].lower() == "cert_keysize":
				if "RSA 2048 BITS" not in row['finding'].upper():
						return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_7(csvfile):
	### Secure Renegotiation (RFC 5746) not supported
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].lower() == "secure_renego":
				if "VULNERABLE" in row['finding'].upper():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_8(csvfile):
	### OCSP Stapling
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "OCSP_STAPLING":
				if "not offered" in row['finding'].lower():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_9(csvfile):
	### Certificate to be expire soon
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "CERT_EXPIRATIONSTATUS":
				if "EXPIRES <" in row['finding'].upper():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def TEST_CASE_10(csvfile):
	### Version Disclosure
	with open(csvfile) as f:
		reader = csv.DictReader(f)
		for row in reader:
			if row["id"].upper() == "BANNER_SERVER":
				if "NO SERVER" not in row['finding'].upper() and "SERVER BANNER IS EMPTY" not in row['finding'].upper():
					return (row["fqdn/ip"].split('/')[1] + " (" + row["port"] + ")")

def make_excel(di,somepath):
	df = pd.DataFrame(di)
	writer = pd.ExcelWriter(somepath, engine='xlsxwriter')
	df.to_excel(writer, sheet_name='Sheet1', index=False)
	writer.save()

	wb = openpyxl.load_workbook(somepath)
	ws = wb.active

path = input("[+] Enter full path for the file: ")
try:
	outpath = input("[+] Enter output file with extension .xlsx: ")
	outpath = os.getcwd() + "/" + outpath
except:
	print("[!] Enter a valid excel filename with .xlsx extension !!!!")

f1 = open(path,'r')

#'echo','N','|',

for ip in f1:
	print("[+] Checks for IP: " + ip.strip())
	subprocess.run([testssl,'--csvfile',ip.strip()+'.csv',ip.strip()], capture_output=True, input="N".encode())
	cip = ip.strip() +".csv"
	TC1.append(TEST_CASE_1(cip))
	TC2.append(TEST_CASE_2(cip))
	TC3.append(TEST_CASE_3(cip))
	TC4.append(TEST_CASE_4(cip))
	TC5.append(TEST_CASE_5(cip))
	TC6.append(TEST_CASE_6(cip))
	TC7.append(TEST_CASE_7(cip))
	TC8.append(TEST_CASE_8(cip))
	TC9.append(TEST_CASE_9(cip))
	TC10.append(TEST_CASE_10(cip))
	subprocess.run(["rm","-rf",cip])

DM['Certificate Expired'] = TC1
DM['SSLv3 POODLE'] = TC2
DM['TLS Server Supports TLS v1.0 and TLS v1.1 and Weak Cipher Algorithms'] = TC3
DM['Secure Client-Initiated Renegotiation'] = TC4
DM['HSTS Policy Not Implemented'] = TC5
DM['SSL/TLS Diffie-Hellman Modulus <= 1024 Bits (Logjam)'] = TC6
DM['Secure Renegotiation (RFC 5746) not supported'] = TC7
DM['OCSP Stapling Not Implemented'] = TC8
DM['Certificate to be expire soon'] = TC9
DM['Version Disclosure'] = TC10

df = pd.DataFrame(DM)

writer = pd.ExcelWriter(outpath, engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1', index=False)
writer.save()

wb = openpyxl.load_workbook(outpath)
ws = wb.active

font_style_1 = Font(color="FF0000")
font_style_2 = Font(color="FFC000")
font_style_3 = Font(color="4472C4")
font_style_4 = Font(color="548235")
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']
d1 = ws['D1']
e1 = ws['E1']
f1 = ws['F1']
g1 = ws['G1']
h1 = ws['H1']
i1 = ws['I1']
j1 = ws['J1']
a1.font = font_style_1
b1.font = font_style_1
c1.font = font_style_2
d1.font = font_style_2
e1.font = font_style_3
f1.font = font_style_3
g1.font = font_style_3
h1.font = font_style_4
i1.font = font_style_4
j1.font = font_style_4
wb.save(outpath)